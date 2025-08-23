# admin_export.py
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side
from django.http import HttpResponse
from django.contrib import admin
from django.utils import timezone
from .models import Lead


def export_leads_to_excel(modeladmin, request, queryset):
    """
    Экспорт выбранных заявок в Excel
    """
    # Создаем новую книгу Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заявки"

    # Заголовки столбцов
    headers = [
        'ID',
        'ФИО',
        'Телефон',
        'Адрес',
        'Регион',
        'Тариф',
        'Цена',
        'Скорость',
        'ТВ',
        'Кинотеатр',
        'Моб. интернет (ГБ)',
        'Минуты',
        'SMS',
        'Статус',
        'Дата создания',
        'Дата обновления',
        'Дата монтажа',
        'Оператор',
        'Примечания'
    ]

    # Стили для заголовков
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Записываем заголовки
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        cell.border = thin_border

    # Заполняем данные
    for row_num, lead in enumerate(queryset, 2):
        ws.cell(row=row_num, column=1, value=lead.id)
        ws.cell(row=row_num, column=2, value=lead.fio)
        ws.cell(row=row_num, column=3, value=lead.phone)
        ws.cell(row=row_num, column=4, value=lead.address)
        ws.cell(row=row_num, column=5, value=str(lead.region))
        ws.cell(row=row_num, column=6, value=lead.tariff.name)
        ws.cell(row=row_num, column=7, value=lead.tariff.price)
        ws.cell(row=row_num, column=8, value=lead.tariff.speed or '')
        ws.cell(row=row_num, column=9, value='Да' if lead.tariff.interactive_tv else 'Нет')
        ws.cell(row=row_num, column=10, value='Да' if lead.tariff.online_cinema else 'Нет')
        ws.cell(row=row_num, column=11, value=lead.tariff.mobile_data)
        ws.cell(row=row_num, column=12, value=lead.tariff.mobile_minutes)
        ws.cell(row=row_num, column=13, value=lead.tariff.mobile_sms)
        ws.cell(row=row_num, column=14, value=lead.get_status_display())
        ws.cell(row=row_num, column=15, value=timezone.localtime(lead.created_at).strftime('%d.%m.%Y %H:%M'))
        ws.cell(row=row_num, column=16,
                value=timezone.localtime(lead.updated_at).strftime('%d.%m.%Y %H:%M') if lead.updated_at else '')
        ws.cell(row=row_num, column=17, value=timezone.localtime(lead.installation_date).strftime(
            '%d.%m.%Y %H:%M') if lead.installation_date else '')
        ws.cell(row=row_num, column=18, value=lead.operator.username if lead.operator else '')
        ws.cell(row=row_num, column=19, value=lead.notes or '')

    # Настраиваем ширину столбцов
    column_widths = [8, 25, 15, 30, 20, 20, 10, 12, 8, 12, 15, 10, 8, 15, 16, 16, 16, 15, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Применяем границы ко всем ячейкам
    for row in ws.iter_rows(min_row=1, max_row=len(queryset) + 1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
            if cell.row != 1:  # Не заголовок
                cell.alignment = Alignment(vertical='center', wrap_text=True)

    # Создаем HTTP response с файлом Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=leads_export_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'

    wb.save(response)
    return response


export_leads_to_excel.short_description = "📊 Экспорт выбранных заявок в Excel"


def export_all_leads_to_excel(modeladmin, request, queryset):
    """
    Экспорт всех заявок в Excel
    """
    # Получаем все заявки
    all_leads = Lead.objects.all()
    return export_leads_to_excel(modeladmin, request, all_leads)


# admin_export.py (дополнительная функция)
def export_filtered_leads(modeladmin, request, queryset):
    """
    Экспорт с фильтрацией по датам через форму
    """
    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')

        leads = Lead.objects.all()
        if start_date:
            leads = leads.filter(created_at__date__gte=start_date)
        if end_date:
            leads = leads.filter(created_at__date__lte=end_date)

        return export_leads_to_excel(modeladmin, request, leads)

    # Показать форму для выбора дат
    from django.shortcuts import render
    return render(request, 'admin/leads_export_form.html')


export_filtered_leads.short_description = "📅 Экспорт заявок по датам"

export_all_leads_to_excel.short_description = "📊 Экспорт всех заявок в Excel"